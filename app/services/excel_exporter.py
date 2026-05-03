from __future__ import annotations

import json
from copy import copy
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import DailyReview, StockOperation


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_review_workbook(review: DailyReview, operations: list[StockOperation]) -> BytesIO:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "每日复盘"
    ws_market = wb.create_sheet("市场数据")
    ws_operations = wb.create_sheet("股票操作")
    ws_ai = wb.create_sheet("AI总结")

    _write_summary(ws_summary, review)
    _write_market(ws_market, review.market_snapshot)
    _write_operations(ws_operations, operations, review.summary)
    _write_ai(ws_ai, review.summary)

    for sheet in wb.worksheets:
        _autosize(sheet)
        sheet.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _write_summary(ws: Any, review: DailyReview) -> None:
    summary = review.summary
    rows = [
        ("交易日", review.trade_date.isoformat()),
        ("模型", review.model_name),
        ("大盘判断", summary.get("market_direction", "")),
        ("涨停家数", summary.get("limit_up_count", 0)),
        ("跌停家数", summary.get("limit_down_count", 0)),
        ("上涨家数", summary.get("rising_count", 0)),
        ("下跌家数", summary.get("falling_count", 0)),
        ("经验和教训", summary.get("lessons", "")),
        ("风险提示", summary.get("risk_notes", "")),
        ("免责声明", summary.get("disclaimer", "")),
    ]
    ws.append(["项目", "内容"])
    for row in rows:
        ws.append(list(row))
    _style_header(ws, 1)
    ws["B8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B9"].alignment = Alignment(wrap_text=True, vertical="top")


def _write_market(ws: Any, market: dict[str, Any]) -> None:
    ws.append(["类型", "名称/项目", "数值1", "数值2", "备注"])
    for item in market.get("indexes", []):
        ws.append(["指数", item.get("name"), item.get("close"), item.get("pct_change"), item.get("date")])
    breadth = market.get("breadth", {})
    for key in ("total_count", "rising_count", "falling_count", "flat_count"):
        ws.append(["涨跌家数", key, breadth.get(key), "", breadth.get("source")])
    limit_stats = market.get("limit_stats", {})
    ws.append(["涨跌停", "limit_up_count", limit_stats.get("limit_up_count"), "", limit_stats.get("source")])
    ws.append(["涨跌停", "limit_down_count", limit_stats.get("limit_down_count"), "", limit_stats.get("source")])
    for item in market.get("hot_sectors", []):
        ws.append(["热点", item.get("sector"), item.get("limit_up_count"), ", ".join(item.get("sample_stocks", [])), ""])
    for warning in market.get("data_quality", {}).get("warnings", []):
        ws.append(["数据质量", "warning", warning, "", ""])
    _style_header(ws, 1)


def _write_operations(ws: Any, operations: list[StockOperation], summary: dict[str, Any]) -> None:
    ws.append(
        [
            "股票代码",
            "股票名称",
            "日期",
            "动作",
            "选股理由",
            "买入理由",
            "持股理由",
            "卖出理由",
            "价格",
            "数量",
            "盈亏",
            "经验教训",
            "AI复盘",
        ]
    )
    reviews = {str(item.get("stock_code")): item for item in summary.get("operation_reviews", [])}
    for operation in operations:
        ai_review = reviews.get(operation.stock_code, {})
        ai_text = "\n".join(
            str(value)
            for value in (
                ai_review.get("selection_reason_review"),
                ai_review.get("buy_reason_review"),
                ai_review.get("hold_reason_review"),
                ai_review.get("sell_reason_review"),
                ai_review.get("profit_loss_review"),
                ai_review.get("lessons"),
            )
            if value
        )
        ws.append(
            [
                operation.stock_code,
                operation.stock_name,
                operation.trade_date.isoformat(),
                operation.action,
                operation.selection_reason,
                operation.buy_reason,
                operation.hold_reason,
                operation.sell_reason,
                float(operation.price) if operation.price is not None else None,
                float(operation.quantity) if operation.quantity is not None else None,
                float(operation.profit_loss) if operation.profit_loss is not None else None,
                operation.lessons,
                ai_text,
            ]
        )
    _style_header(ws, 1)


def _write_ai(ws: Any, summary: dict[str, Any]) -> None:
    ws.append(["字段", "内容"])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False, indent=2) if isinstance(value, (dict, list)) else value])
    _style_header(ws, 1)
    for row in ws.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def _style_header(ws: Any, row_number: int) -> None:
    for cell in ws[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws: Any) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(60, len(value) + 2))
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width
