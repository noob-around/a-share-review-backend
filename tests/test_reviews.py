from io import BytesIO

from openpyxl import load_workbook


def test_create_review_and_download_excel(client, auth_headers):
    operation_payload = {
        "stock_code": "600519",
        "stock_name": "贵州茅台",
        "trade_date": "2026-05-03",
        "action": "买入",
        "buy_reason": "回踩企稳。",
        "price": "1680",
        "quantity": "100",
        "profit_loss": "100",
        "lessons": "按计划执行。",
    }
    response = client.post("/api/operations", json=operation_payload, headers=auth_headers)
    assert response.status_code == 201

    response = client.post(
        "/api/reviews/daily",
        json={"trade_date": "2026-05-03", "refresh": True},
        headers=auth_headers,
    )
    assert response.status_code == 201, response.text
    review = response.json()
    assert review["trade_date"] == "2026-05-03"
    assert review["summary"]["market_direction"] == "大盘整体上涨"
    assert review["summary"]["limit_up_count"] == 2

    response = client.get(f"/api/reviews/{review['id']}/excel", headers=auth_headers)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["每日复盘", "市场数据", "股票操作", "AI总结"]
    assert wb["每日复盘"]["A1"].value == "项目"
    assert wb["股票操作"]["A2"].value == "600519"
    assert wb["AI总结"]["A2"].value == "market_direction"


def test_existing_review_returned_without_refresh(client, auth_headers):
    payload = {"trade_date": "2026-05-03", "refresh": True}
    first = client.post("/api/reviews/daily", json=payload, headers=auth_headers)
    assert first.status_code == 201
    second = client.post("/api/reviews/daily", json={"trade_date": "2026-05-03"}, headers=auth_headers)
    assert second.status_code == 201
    assert second.json()["id"] == first.json()["id"]
